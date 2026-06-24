from flask import Blueprint, jsonify, request, render_template, redirect, url_for, flash,session,send_file
from datetime import datetime,timedelta,timezone,date
import calendar
from app import db
from app.models.farmacia import Medicamento,GrupoTerapeutico,MaterialFamilia, AsignacionReceta,MovimientoAlmacenFarmacia,InventarioAlmacen,InventarioFarmacia,EntradaAlmacen,TransferenciaSaliente, SalidaFarmacia,TransferenciaEntrante, Empleado,BitacoraMovimiento
from app.models.personal import Usuario, Empleado
from flask_login import current_user
from app.utils.helpers import roles_required
from sqlalchemy.orm import joinedload, selectinload
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import func
import os
import pandas as pd

import io

from flask import make_response, request  # Añadido request para capturar el parámetro 'q'

# Importaciones requeridas para construir el PDF con ReportLab
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas


bp = Blueprint('farmacia', __name__, template_folder='templates/farmacia')

# Página principal de farmacia
@bp.route('/')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def index():
    # 1. Capturar el criterio de búsqueda desde el formulario de la plantilla
    query = request.args.get('q', '').strip()
    
    # 2. Base de la consulta: Traer los medicamentos usando tu modelo real
    medicamentos_query = Medicamento.query
    
    # 3. Aplicar filtro de búsqueda usando tus campos exactos
    if query:
        medicamentos_query = medicamentos_query.filter(
            (Medicamento.clave.ilike(f'%{query}%')) |        
            (Medicamento.principio_activo.ilike(f'%{query}%')) 
        )
    
    # 🌟 CORRECCIÓN: Limitamos la consulta a solo 5 registros antes de traer los resultados con .all()
    lista_medicamentos = medicamentos_query.limit(5).all()
    inventario = []
    
    # 4. Calcular Entradas, Salidas Totales y Existencias usando 'id_medicamento'
    for med in lista_medicamentos:
        # Sumar total de entradas desde la tabla EntradaAlmacen
        total_entradas = db.session.query(func.sum(EntradaAlmacen.cantidad))\
            .filter(EntradaAlmacen.id_medicamento == med.id_medicamento).scalar() or 0
            
        # Sumar TODAS las salidas desde SalidaFarmacia
        total_salidas = db.session.query(func.sum(SalidaFarmacia.cantidad))\
            .filter(SalidaFarmacia.id_medicamento == med.id_medicamento).scalar() or 0
            
        # Operación matemática del Stock real en almacén
        existencia = total_entradas - total_salidas
        
        # Guardar los datos estructurados mapeando tus columnas reales
        inventario.append({
            'id_medicamento': med.id_medicamento,
            'clave': med.clave,
            'principio_activo': med.principio_activo or 'Sin descripción',
            'presentacion': med.presentacion or '',
            'concentracion': med.concentracion or '',
            'unidad': med.unidad or '',
            'stock_minimo': med.stock_minimo,
            'total_entradas': total_entradas,
            'total_salidas': total_salidas,
            'existencia': existencia
        })
        
    # 5. Renderizar la plantilla enviando las variables requeridas
    return render_template(
        'farmacia/index.html', 
        inventario=inventario, 
        query=query
    )

@bp.context_processor
def inject_today():
    """Inyecta la FECHA PURA (date) sin hora para evitar el error de comparación"""
    return dict(today=datetime.now().date()) # 🌟 Añadimos .date() al final

#)))))))))))))))))))))))))))))(((((((((((((((((*********************************** otras salidas es colectivo , oc99
@bp.route('/salidas/otras', methods=['GET', 'POST'])
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def otras_salidas():
    # 🟢 SI EL USUARIO PRESIONA "PROCESAR SALIDA TOTAL" (MANUAL)
    if request.method == 'POST':
        # 1. Capturamos los datos generales de la interfaz
        tipo_salida = request.form.get('tipo_salida')
        entidad_destino = request.form.get('entidad_destino')   # Campo "Servicio Destino Interno"
        documento_soporte = request.form.get('documento_soporte') # Campo "Número de Vale Colectivo"

        # 2. ✨ BLINDAJE ESTRICTO DEL ENUM CONTRA 'OTRA_SALIDA' O 'NONE'
        if not tipo_salida or tipo_salida == 'OTRA_SALIDA':
            destino_limpio = entidad_destino.upper() if entidad_destino else ''
            
            # Deducimos el Enum correcto según lo que el usuario escribió en la interfaz
            if 'CEYE' in destino_limpio or 'PISO' in destino_limpio or 'VALE' in destino_limpio:
                tipo_salida_final = 'COLECTIVO'        # Salva tu inserción actual de CEYE
            elif 'CADUC' in destino_limpio:
                tipo_salida_final = 'BAJA_CADUCIDAD'
            elif 'EXTRAV' in destino_limpio or 'FALTANTE' in destino_limpio:
                tipo_salida_final = 'BAJA_EXTRAVIO'
            elif 'TRASLADO' in destino_limpio or 'HOSPITAL' in destino_limpio:
                tipo_salida_final = 'TRASLADO_UNIDAD'
            else:
                tipo_salida_final = 'COLECTIVO'        # Respaldo seguro que existe en tu Enum
        else:
            tipo_salida_final = tipo_salida

        # 3. Capturamos las listas dinámicas del abanico de filas de la tabla
        medicamentos_ids = request.form.getlist('id_medicamento[]')
        lotes = request.form.getlist('lote[]')
        cantidades = request.form.getlist('cantidad[]')

        if not medicamentos_ids:
            flash('No se agregaron medicamentos al abanico.', 'danger')
            return redirect(url_for('farmacia.otras_salidas'))

        registros_procesados = 0
        errores = []

        try:
            # Usamos un bloque nested por si ocurre un error de stock poder hacer rollback
            with db.session.begin_nested():
                for i in range(len(medicamentos_ids)):
                    id_med = int(medicamentos_ids[i])
                    cant_solicitada = int(cantidades[i])
                    lote_txt = lotes[i].strip()

                    # 4. Buscamos el lote específico seleccionado en el inventario
                    lote_inventario = InventarioFarmacia.query.filter_by(
                        id_medicamento=id_med, 
                        lote=lote_txt
                    ).first()

                    # 5. VALIDACIÓN: Verificar si el lote existe y tiene stock suficiente
                    if not lote_inventario:
                        errores.append(f"El lote '{lote_txt}' no existe en el inventario.")
                        continue
                        
                    if lote_inventario.cantidad < cant_solicitada:
                        errores.append(f"Stock insuficiente en lote '{lote_txt}'. Disponible: {lote_inventario.cantidad}, Solicitado: {cant_solicitada}")
                        continue

                    # 6. ✨ OPERACIÓN CRUCIAL: Descontar la cantidad directamente del anaquel
                    lote_inventario.cantidad -= cant_solicitada
                    fecha_venc = lote_inventario.fecha_vencimiento

                    # 7. Creamos el registro de salida garantizando un valor del Enum válido para Postgres
                    nueva_salida = SalidaFarmacia(
                        id_medicamento=id_med,
                        cantidad=cant_solicitada,
                        lote=lote_txt,
                        fecha_vencimiento=fecha_venc,
                        fecha_salida=datetime.now(timezone.utc),
                        id_usuario=current_user.id_usuario,
                        tipo_salida=tipo_salida_final,  # 'COLECTIVO', 'BAJA_CADUCIDAD', ETC.
                        entidad_destino=entidad_destino,
                        documento_soporte=documento_soporte
                    )
                    db.session.add(nueva_salida)

                    # 8. Creamos la bitácora de movimiento correspondientemente
                    nueva_bitacora = BitacoraMovimiento(
                        id_medicamento=id_med,
                        id_usuario=current_user.id_usuario,
                        fecha_hora=datetime.now(timezone.utc),
                        movimiento=f"SALIDA_MANUAL_{tipo_salida_final} | Cantidad: {cant_solicitada} | Lote: {lote_txt}"
                    )
                    db.session.add(nueva_bitacora)
                    registros_procesados += 1

            # Si hubo errores de stock, cancelamos el proceso interno y avisamos al usuario
            if errores:
                db.session.rollback()
                for err in errores:
                    flash(err, 'danger')
                return redirect(url_for('farmacia.otras_salidas'))

            # Si todo el bucle del abanico manual termina bien, guardamos de manera definitiva
            db.session.commit()
            flash(f'¡Éxito! Se registró la salida manual de {registros_procesados} insumos y se afectó el inventario.', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'Error al procesar la salida especial de farmacia: {str(e)}', 'danger')

        return redirect(url_for('farmacia.otras_salidas'))

    # 🔵 SI EL NAVEGADOR ENTRA POR PRIMERA VEZ (MÉTODO GET)
    # Coloca aquí tus consultas normales si necesitas mandar catálogos al HTML
    return render_template('farmacia/otras_salidas.html')



# 🌟 ENDPOINT AJAX 1: Obtiene lotes vigentes de farmacia para el renglón de la tabla
@bp.route('/api/lotes_farmacia/<int:id_medicamento>')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def api_lotes_farmacia(id_medicamento):
    try:
        # 1. Forzar una consulta limpia usando db.session
        lotes = db.session.query(InventarioFarmacia).filter(
            InventarioFarmacia.id_medicamento == id_medicamento,
            InventarioFarmacia.cantidad > 0
        ).all()
        
        resultados = []
        for l in lotes:
            # 2. Formatear la fecha de vencimiento sin que rompa si viene como string o vacía
            fecha_str = 'S/V'
            if l.fecha_vencimiento:
                if isinstance(l.fecha_vencimiento, (date, datetime)):
                    fecha_str = l.fecha_vencimiento.strftime('%d/%m/%Y')
                else:
                    fecha_str = str(l.fecha_vencimiento)

            resultados.append({
                'lote': l.lote,
                'cantidad': l.cantidad,
                'fecha_vencimiento': fecha_str
            })
            
        # 3. Responder de forma explícita con el JSON listo
        return jsonify(resultados)

    except Exception as e:
        # 🌟 Esto va a imprimir el error exacto en tu terminal negra de Flask para saber qué falla
        print("\n" + "="*50)
        print(f"❌ ERROR CRÍTICO EN API LOTES: {str(e)}")
        print("="*50 + "\n")
        return jsonify({'error': str(e)}), 500

# 🌟 ENDPOINT AJAX 2: Para el buscador masivo del Select2
@bp.route('/api/buscar_insumos_select2')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def buscar_insumos_select2():
    q = request.args.get('q', '').strip()
    insumos = Medicamento.query.filter(
        (Medicamento.clave.ilike(f"%{q}%")) | 
        (Medicamento.principio_activo.ilike(f"%{q}%"))
    ).limit(20).all()
    
    resultados = [{'id': m.id_medicamento, 'text': f"{m.clave} - {m.principio_activo} ({m.presentacion or ''} {m.via_administracion or ''} {m.concentracion or ''} )"} for m in insumos]
    return jsonify(resultados)



@bp.route('/salidas/carga_masiva_0c99', methods=['POST'])
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def carga_masiva_0c99():
    if 'archivo_excel' not in request.files:
        flash('No se seleccionó ningún archivo', 'danger')
        return redirect(url_for('farmacia.otras_salidas'))
        
    archivo = request.files['archivo_excel']
    if archivo.filename == '':
        flash('Archivo no válido', 'danger')
        return redirect(url_for('farmacia.otras_salidas'))

    if archivo and (archivo.filename.endswith('.xlsx') or archivo.filename.endswith('.xls')):
        try:
            df = pd.read_excel(archivo)
            df.columns = [str(col).strip().upper() for col in df.columns]
            
            if 'CLAVE' not in df.columns:
                flash('El archivo debe contener la columna "CLAVE".', 'danger')
                return redirect(url_for('farmacia.otras_salidas'))

            # 🌟 NUEVA LÓGICA: Determinar qué columnas numéricas del IMSS se van a sumar
            modo_filtro = request.form.get('modo_filtro', 'TODO')
            columnas_a_sumar = []

            if modo_filtro == 'UN_DIA':
                dia = request.form.get('dia_unico', '').strip()
                if dia and dia in df.columns:
                    columnas_a_sumar = [dia]
                else:
                    flash(f'La columna del día {dia} no se encuentra en el archivo subido.', 'danger')
                    return redirect(url_for('farmacia.otras_salidas'))
                    
            elif modo_filtro == 'RANGO_DIAS':
                try:
                    inicio = int(request.form.get('dia_inicio', 1))
                    fin = int(request.form.get('dia_fin', 31))
                    # Crear lista de strings con el rango de días válidos que existan en las columnas
                    columnas_a_sumar = [str(d) for d in range(inicio, fin + 1) if str(d) in df.columns]
                except ValueError:
                    flash('Rango de días no válido.', 'danger')
                    return redirect(url_for('farmacia.otras_salidas'))
            else:
                # Si es TODO, buscamos explícitamente la columna 'TOTAL' de la matriz
                if 'TOTAL' in df.columns:
                    columnas_a_sumar = ['TOTAL']
                else:
                    flash('No se encontró la columna "TOTAL" para procesar el mes completo.', 'danger')
                    return redirect(url_for('farmacia.otras_salidas'))

            registros_procesados = 0
            errores = []
            destino = request.form.get('excel_entidad_destino', 'CONSUMO DIRECTO OC99').upper()
            soporte = request.form.get('excel_documento_soporte', 'S/D').upper()

            with db.session.begin_nested():
                for index, fila in df.iterrows():
                    clave_cruda = str(fila['CLAVE']).strip()
                    if not clave_cruda or clave_cruda.lower() == 'nan':
                        continue

                    # 🌟 MATEMÁTICA EN CALIENTE: Sumamos únicamente las celdas de los días elegidos
                    cantidad_calculada = 0
                    for col in columnas_a_sumar:
                        try:
                            val_celda = float(fila[col])
                            if not pd.isna(val_celda):
                                cantidad_calculada += int(val_celda)
                        except (ValueError, TypeError):
                            continue

                    # Si para este rango el medicamento no tuvo salidas, pasamos de largo
                    if cantidad_calculada <= 0:
                        continue

                    med = Medicamento.query.filter_by(clave=clave_cruda).first()
                    if not med:
                        errores.append(f"Fila {index+2}: La clave '{clave_cruda}' no existe en el catálogo.")
                        continue

                    # Buscar stock bajo orden PEPS
                    lotes_disponibles = db.session.query(InventarioFarmacia).filter(
                        InventarioFarmacia.id_medicamento == med.id_medicamento,
                        InventarioFarmacia.cantidad > 0
                    ).order_by(InventarioFarmacia.fecha_vencimiento.asc()).all()

                    total_farmacia = sum(l.cantidad for l in lotes_disponibles)

                    if total_farmacia < cantidad_calculada:
                        errores.append(f"Clave {med.clave}: Stock insuficiente. Solicitado ({modo_filtro}): {cantidad_calculada}, En Anaquel: {total_farmacia}")
                        continue

                    # Aplicar descuentos lote por lote
                    por_descontar = cantidad_calculada
                    for lote_farmacia in lotes_disponibles:
                        if por_descontar <= 0:
                            break

                        if lote_farmacia.cantidad >= por_descontar:
                            cant_desde_este_lote = por_descontar
                            lote_farmacia.cantidad -= por_descontar
                            por_descontar = 0
                        else:
                            cant_desde_este_lote = lote_farmacia.cantidad
                            por_descontar -= lote_farmacia.cantidad
                            lote_farmacia.cantidad = 0

                        # Registrar afectación
                        nueva_salida = SalidaFarmacia(
                            id_medicamento=med.id_medicamento,
                            cantidad=cant_desde_este_lote, # Ajusta a tu campo 'cantidad'
                            lote=lote_farmacia.lote,
                            fecha_vencimiento=lote_farmacia.fecha_vencimiento,
                            fecha_salida=datetime.now(timezone.utc),
                            id_usuario=current_user.id_usuario,
                            tipo_salida='BAJA_0C99'
                        )
                        db.session.add(nueva_salida)

                        nueva_bitacora = BitacoraMovimiento(
    						id_medicamento=med.id_medicamento,
    						id_usuario=current_user.id_usuario,
    						fecha_hora=datetime.now(timezone.utc),
    						# Guardamos la cantidad como texto aquí adentro
    						movimiento=f"SALIDA_BAJA_0C99 | Cantidad: {cant_desde_este_lote} | Lote: {lote_farmacia.lote}"
						)
                        db.session.add(nueva_bitacora)
                        registros_procesados += 1

            if errores:
                db.session.rollback()
                return render_template('farmacia/otras_salidas.html', errores_excel=errores)

            db.session.commit()
            flash(f'¡Éxito! Procesado en modo {modo_filtro}. Se afectaron {registros_procesados} lotes en el inventario.', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al procesar el documento Excel: {str(e)}', 'danger')
            
    return redirect(url_for('farmacia.otras_salidas'))



#}||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||

# Registrar salida
@bp.route('/salida/<int:id_medicamento>', methods=['GET', 'POST'])
@roles_required([ 'UsuarioAdministrativo', 'Administrador'])
def registrar_salida(id_medicamento):
    medicamento = Medicamento.query.get_or_404(id_medicamento)
    medicos = Usuario.query.all()

    if request.method == 'POST':
        cantidad = int(request.form['cantidad'])
        id_medico = int(request.form['id_medico'])

        asignacion = AsignacionReceta.query.filter_by(
            id_medico=id_medico
        ).order_by(AsignacionReceta.id_asignacion.desc()).first()

        if not asignacion:
            flash("⚠️ El médico no tiene bloque asignado", "danger")
            return redirect(url_for('farmacia.registrar_salida', id_medicamento=id_medicamento))

        folio = asignacion.siguiente_folio()
        if not folio:
            flash("⚠️ El bloque asignado está agotado", "danger")
            return redirect(url_for('farmacia.registrar_salida', id_medicamento=id_medicamento))

        salida = SalidaFarmacia(
            id_medicamento=id_medicamento,
            cantidad=cantidad,
            fecha_salida=datetime.utcnow(),
            id_usuario=id_medico,
            folio_receta=folio
        )

        db.session.add(salida)
        db.session.commit()

        flash(f"✅ Salida registrada con folio {folio}", "success")
        return redirect(url_for('farmacia.listar_salidas'))

    return render_template('farmacia/salida.html', medicamento=medicamento, medicos=medicos)


# Listar todas las salidas 
@bp.route('/salidas')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def listar_salidas():
    page = request.args.get('page', 1, type=int)
    per_page = 20

       

    # Consulta maestra optimizada para SQLAlchemy 2.0 usando atributos de clase directos
    query = (
        SalidaFarmacia.query
        .options(
            joinedload(SalidaFarmacia.medicamento),
            joinedload(SalidaFarmacia.receta),
            joinedload(SalidaFarmacia.usuario).joinedload(Usuario.empleado)
        )
        .order_by(SalidaFarmacia.fecha_salida.desc())
    )

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    salidas = pagination.items

    return render_template(
        'farmacia/listar_salidas.html', 
        salidas=salidas, 
        pagination=pagination
    )

#)))))))))) exportar todas las salidas a traves de un 0c99 excel
@bp.route('/descargar_salidas_oc99')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def descargar_salidas_oc99():
    from datetime import datetime
    import calendar
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import request, send_file
    # Asegúrate de importar tus modelos db, Medicamento y Salida aquí si no son globales

    ahora = datetime.now()
    anio = request.args.get('anio', ahora.year, type=int)
    mes = request.args.get('mes', ahora.month, type=int)
    dias_mes = calendar.monthrange(anio, mes)[1]

    # 1. Obtener datos
    medicamentos = db.session.query(Medicamento).order_by(Medicamento.clave).all()
    # Filtramos las salidas por el año y mes seleccionados
    salidas = db.session.query(SalidaFarmacia).filter(
        db.extract('year', SalidaFarmacia.fecha_salida) == anio,
        db.extract('month', SalidaFarmacia.fecha_salida) == mes
    ).all()

    # 2. Matriz de datos y Totales por día
    matriz = {}
    totales_piezas_dia = {d: 0 for d in range(1, dias_mes + 1)}
    claves_por_dia = {d: set() for d in range(1, dias_mes + 1)}

    for med in medicamentos:
        nombre_completo = f"{med.principio_activo or ''} - {med.presentacion or ''}".strip().upper()
        matriz[med.id_medicamento] = {
            'clave': med.clave,
            'nombre': nombre_completo,
            'dias': {d: "" for d in range(1, dias_mes + 1)}  # Inicializado en vacío
        }

    for s in salidas:
        if s.fecha_salida:
            dia = s.fecha_salida.day
            if s.id_medicamento in matriz:
                # Recuperar valor previo controlando si está vacío
                valor_actual = matriz[s.id_medicamento]['dias'][dia]
                cantidad_previa = valor_actual if isinstance(valor_actual, int) else 0
                
                matriz[s.id_medicamento]['dias'][dia] = cantidad_previa + s.cantidad
                totales_piezas_dia[dia] += s.cantidad
                if s.cantidad > 0:
                    claves_por_dia[dia].add(s.id_medicamento)

    # 3. Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "OC99 Salidas"

    # --- CUADRO DE RESUMEN SUPERIOR ---
    font_bold = Font(bold=True)
    fill_resumen = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Fila de TOTAL CLAVES (Vacío si es 0)
    fila_claves = ["TOTAL CLAVES:", ""] + [len(claves_por_dia[d]) if len(claves_por_dia[d]) > 0 else "" for d in range(1, dias_mes + 1)]
    ws.append(fila_claves)
    
    # Fila de TOTAL PIEZAS (Vacío si es 0)
    fila_piezas = ["TOTAL PIEZAS:", ""] + [totales_piezas_dia[d] if totales_piezas_dia[d] > 0 else "" for d in range(1, dias_mes + 1)]
    ws.append(fila_piezas)

    # Estilo para el resumen
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = font_bold
            cell.fill = fill_resumen
            cell.alignment = Alignment(horizontal="center")

    ws.append([]) # Espacio en blanco

    # 4. ENCABEZADOS DE LA TABLA
    headers = ["CLAVE", "MEDICAMENTO"] + [str(d) for d in range(1, dias_mes + 1)]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    for cell in ws[4]: # La fila 4 es el encabezado
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 5. DATOS DE MEDICAMENTOS
    for med in matriz.values():
        fila = [med['clave'], med['nombre']] + [med['dias'][d] for d in range(1, dias_mes + 1)]
        ws.append(fila)

    # 6. AJUSTES FINALES
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 60
    
    # Ajustar ancho de columnas de días
    for col_idx in range(3, 3 + dias_mes):
        ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = 5

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name=f"OC99_SALIDAS_{mes:02d}_{anio}.xlsx", as_attachment=True)




@bp.route('/reporte_medicamentos')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def reporte_medicamentos():
    query = request.args.get('q', '').strip()
    
    # 1. Consulta limpia uniendo únicamente las tablas directas de tu modelo
    salidas_query = db.session.query(SalidaFarmacia)\
        .join(Usuario, SalidaFarmacia.id_usuario == Usuario.id_usuario)\
        .join(Medicamento, SalidaFarmacia.id_medicamento == Medicamento.id_medicamento)

    # 2. Filtros del buscador por medicamento o por el usuario que despachó
    if query:
        salidas_query = salidas_query.filter(
            (Medicamento.clave.ilike(f'%{query}%')) |
            (Medicamento.principio_activo.ilike(f'%{query}%')) |
            (Usuario.usuario.ilike(f'%{query}%'))
        )

    # 3. Ordenar por la fecha de salida más reciente
    salidas = salidas_query.order_by(SalidaFarmacia.fecha_salida.desc()).all()
    
    return render_template('farmacia/reporte_medicamentos.html', salidas=salidas, query=query)

#)))))))))))))))))))))))))))))))))) medicamento
# Listar medicamentos
@bp.route('/medicamentos')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def listar_medicamentos():
    query = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    
    # Base de la consulta ordenada
    medicamentos_query = Medicamento.query.order_by(Medicamento.principio_activo.asc())

    if query:
        # Filtro con búsqueda parcial
        medicamentos_query = medicamentos_query.filter(
            (Medicamento.clave.ilike(f'%{query}%')) |
            (Medicamento.principio_activo.ilike(f'%{query}%'))
        )
    
    # Paginación (ejemplo: 10 por página)
    pagination = medicamentos_query.paginate(page=page, per_page=10, error_out=False)
    medicamentos = pagination.items

    return render_template(
        'farmacia/listar_medicamentos.html', 
        medicamentos=medicamentos, 
        query=query,
        pagination=pagination
    )


# Nuevo medicamento
@bp.route('/medicamentos/nuevo', methods=['GET', 'POST'])
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def nuevo_medicamento():
    if request.method == 'POST':
        # 1. Captura y Limpieza (eliminar espacios accidentales)
        clave = request.form.get('clave', '').strip().upper()
        principio = request.form.get('principio_activo', '').strip().upper()
        presentacion = request.form.get('presentacion', '').strip().upper()

        # 2. VALIDACIÓN: Campos obligatorios vacíos o puros espacios
        if not clave or not principio or not presentacion:
            flash("❌ La clave, el principio activo y la presentación son campos obligatorios.", "warning")
            return redirect(url_for('farmacia.nuevo_medicamento'))

        # 3. VALIDACIÓN: Clave única
        if Medicamento.query.filter_by(clave=clave).first():
            flash(f"⚠️ La clave '{clave}' ya existe. Intenta con otra.", "danger")
            return redirect(url_for('farmacia.nuevo_medicamento'))

        # 4. VALIDACIÓN: Datos numéricos
        try:
            s_min = int(request.form.get('stock_minimo') or 0)
            s_max = int(request.form.get('stock_maximo') or 0)
            cpm_val = float(request.form.get('cpm') or 0.0)
        except ValueError:
            flash("❌ Los stocks deben ser números enteros y el CPM decimal.", "danger")
            return redirect(url_for('farmacia.nuevo_medicamento'))

        # 5. VALIDACIÓN: Lógica de inventario
        if s_min > s_max:
            flash(f"⚠️ Stock Mínimo ({s_min}) no puede ser mayor al Máximo ({s_max}).", "warning")
            return redirect(url_for('farmacia.nuevo_medicamento'))

        try:
            # 6. CREACIÓN DEL REGISTRO
            nuevo_med = Medicamento(
                clave=clave,
                principio_activo=principio,
                presentacion=presentacion,
                via_administracion=request.form.get('via_administracion', '').strip().upper(),
                concentracion=request.form.get('concentracion', '').strip().upper(),
                unidad=request.form.get('unidad', '').strip().upper(),
                
                # Checkboxes
                es_kit_basico='es_kit_basico' in request.form,
                es_180_claves='es_180_claves' in request.form,
                es_general='es_general' in request.form,
                
                stock_minimo=s_min,
                stock_maximo=s_max,
                cpm=cpm_val,
                nivel_movimiento=request.form.get('nivel_movimiento', 'Nulo').capitalize()
            )
            
            db.session.add(nuevo_med)
            db.session.commit()
            flash("✅ Medicamento registrado correctamente", "success")
            return redirect(url_for('farmacia.listar_medicamentos'))

        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error de base de datos: {str(e)}", "danger")
            return redirect(url_for('farmacia.nuevo_medicamento'))

    return render_template('farmacia/nuevo_medicamento.html')

# Editar medicamento
@bp.route('/medicamentos/editar/<int:id_medicamento>', methods=['GET', 'POST'])
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def editar_medicamento(id_medicamento):
    medicamento = Medicamento.query.get_or_404(id_medicamento)

    if request.method == 'POST':
        # 1. Captura y Limpieza inicial
        nueva_clave = request.form.get('clave', '').strip().upper()
        principio = request.form.get('principio_activo', '').strip().upper()
        presentacion = request.form.get('presentacion', '').strip().upper()

        # --- VALIDACIÓN DE CAMPOS VACÍOS ---
        if not nueva_clave or not principio or not presentacion:
            flash("❌ La clave, el principio activo y la presentación no pueden estar vacíos.", "warning")
            return redirect(url_for('farmacia.editar_medicamento', id_medicamento=id_medicamento))

        # 2. Validación de Stocks (Números)
        try:
            s_min = int(request.form.get('stock_minimo') or 0)
            s_max = int(request.form.get('stock_maximo') or 0)
            cpm_val = float(request.form.get('cpm') or 0.0)
        except ValueError:
            flash("❌ Los valores de stock y CPM deben ser numéricos.", "danger")
            return redirect(url_for('farmacia.editar_medicamento', id_medicamento=id_medicamento))

        if s_min > s_max:
            flash(f"⚠️ El Stock Mínimo ({s_min}) no puede ser mayor al Máximo ({s_max}).", "warning")
            return redirect(url_for('farmacia.editar_medicamento', id_medicamento=id_medicamento))

        # 3. Validación de Clave Única
        med_existente = Medicamento.query.filter(
            Medicamento.clave == nueva_clave, 
            Medicamento.id_medicamento != id_medicamento
        ).first()
        
        if med_existente:
            flash(f"⚠️ La clave '{nueva_clave}' ya pertenece a otro medicamento.", "danger")
            return redirect(url_for('farmacia.editar_medicamento', id_medicamento=id_medicamento))

        try:
            # 4. Actualización
            medicamento.clave = nueva_clave
            medicamento.principio_activo = principio
            medicamento.presentacion = presentacion
            medicamento.via_administracion = request.form.get('via_administracion', '').strip().upper()
            medicamento.concentracion = request.form.get('concentracion', '').strip().upper()
            medicamento.unidad = request.form.get('unidad', '').strip().upper()
            
            medicamento.es_kit_basico = 'es_kit_basico' in request.form
            medicamento.es_180_claves = 'es_180_claves' in request.form
            medicamento.es_general = 'es_general' in request.form
            
            medicamento.stock_minimo = s_min
            medicamento.stock_maximo = s_max
            medicamento.cpm = cpm_val

            # CORRECCIÓN DEL ENUM PARA POSTGRESQL
            nivel_form = request.form.get('nivel_movimiento', 'Nulo')
            medicamento.nivel_movimiento = nivel_form.capitalize() 

            db.session.commit()
            flash("✅ Medicamento actualizado correctamente", "success")
            return redirect(url_for('farmacia.listar_medicamentos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error crítico de base de datos: {str(e)}", "danger")

    return render_template('farmacia/editar_medicamento.html', medicamento=medicamento)



# Eliminar medicamento (con protección)
@bp.route('/medicamentos/eliminar/<int:id_medicamento>', methods=['POST'])
@roles_required(['Administrador'])
def eliminar_medicamento(id_medicamento):
    medicamento = Medicamento.query.get_or_404(id_medicamento)
    
    # Verificamos si tiene inventario antes de borrar para evitar error de SQL
    if medicamento.inventario_almacen or medicamento.inventario_farmacia:
        flash("❌ No se puede eliminar: El medicamento tiene existencias en inventario.", "danger")
        return redirect(url_for('farmacia.listar_medicamentos'))
        
    db.session.delete(medicamento)
    db.session.commit()
    flash("✅ Medicamento eliminado correctamente", "success")
    return redirect(url_for('farmacia.listar_medicamentos'))



#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++Grupo Terapeutico +++++++++++++++++++++++++++++++++++++++++
@bp.route('/GrupoTerapeutico')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def lista_grupos_terapeuticos():
    # 1. CONSULTA PARA MEDICAMENTOS (Grupos 1 al 23)
    # Filtramos para asegurarnos de contar solo registros que inicien con 010 o 040
    resultados_medicamentos = db.session.query(
        GrupoTerapeutico.grupo_id,
        GrupoTerapeutico.nombre_grupo,
        GrupoTerapeutico.rango_inicio,
        GrupoTerapeutico.rango_fin,
        db.func.count(Medicamento.clave) # Usamos tu columna real 'clave'
    ).join(
        Medicamento, 
        Medicamento.grupo_id == GrupoTerapeutico.grupo_id,
        isouter=True
    ).group_by(
        GrupoTerapeutico.grupo_id,
        GrupoTerapeutico.nombre_grupo,
        GrupoTerapeutico.rango_inicio,
        GrupoTerapeutico.rango_fin
    ).order_by(
        GrupoTerapeutico.grupo_id.asc()
    ).all()

    # 2. CONSULTA PARA MATERIAL DE CURACIÓN Y REACTIVOS (Las 74 familias)
    # Asumiendo que tu modelo de SQLAlchemy para la nueva tabla se llama 'MaterialFamilia'
    resultados_materiales = db.session.query(
        MaterialFamilia.familia_id,
        MaterialFamilia.nombre_familia,
        db.func.count(Medicamento.clave)
    ).join(
        Medicamento,
        Medicamento.material_familia_id == MaterialFamilia.familia_id,
        isouter=True
    ).group_by(
        MaterialFamilia.familia_id,
        MaterialFamilia.nombre_familia
    ).order_by(
        MaterialFamilia.familia_id.asc()
    ).all()
    
    etiquetas_catalogo = {
        '010': '010 - Medicamentos',
        '040': '040 - Psicotrópicos',
        '060': '060 - Mat. Curación',
        '070': '070 - Laboratorio',
        '080': '080 - Radiología'
    }

    datos_procesados = []

    # 3. PROCESAR MEDICAMENTOS (Identifica si es 010 o 040 de forma dinámica)
    for grupo_id, nombre, inicio, fin, total in resultados_medicamentos:
        # Si el grupo es 19 (Psiquiatría) y manejas Naloxona, puede ser 040, si no 010
        tipo = '040' if grupo_id == 19 else '010' 
        
        datos_procesados.append({
            'id': grupo_id,
            'nombre': nombre,
            'tipo': tipo,
            'tipo_label': etiquetas_catalogo.get(tipo),
            'inicio': inicio,
            'fin': fin,
            'total': total
        })

    # 4. PROCESAR MATERIALES, LABORATORIO Y RADIOLOGÍA
    for familia_id, nombre, total in resultados_materiales:
        # Detectamos dinámicamente el tipo leyendo el inicio de la familia
        # Las claves 060, 070 y 080 se identifican por sus rangos cargados
        if familia_id in ['707']: 
            tipo = '070' # Laboratorio / Placas dentales
        elif familia_id in ['018', '855', '889']: 
            tipo = '080' # Radiología y servicios / Tiras reactivas
        else: 
            tipo = '060' # Material de curación estándar (Agujas, gasas, etc.)

        datos_procesados.append({
            'id': familia_id,
            'nombre': nombre,
            'tipo': tipo,
            'tipo_label': etiquetas_catalogo.get(tipo),
            'inicio': 'N/A', # El material de curación no usa rangos
            'fin': 'N/A',
            'total': total
        })

    return render_template('farmacia/grupoterapeutico.html', grupos=datos_procesados)


@bp.route('/GrupoTerapeutico/<int:grupo_id>')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def detalle_GrupoTerapeutico(grupo_id):
    grupo = db.get_or_404(GrupoTerapeutico, grupo_id)
    
    medicamentos_con_stock = []
    for med in grupo.medicamentos:
        # 1. Sumamos existencias reales de almacén y farmacia
        stock_almacen = sum(inv.cantidad for inv in med.inventario_almacen if inv.cantidad)
        stock_farmacia = sum(inv.cantidad for inv in med.inventario_farmacia if inv.cantidad)
        stock_total = stock_almacen + stock_farmacia
        
        # 2. Extraemos los lotes únicos registrados en ambos inventarios
        lotes = set()
        for inv in med.inventario_almacen:
            if getattr(inv, 'lote', None): lotes.add(inv.lote)
        for inv in med.inventario_farmacia:
            if getattr(inv, 'lote', None): lotes.add(inv.lote)
        lotes_str = ", ".join(lotes) if lotes else "N/A"
        
        # 3. Recolectamos las fechas de vencimiento de ambas tablas
        caducidades = []
        for inv in med.inventario_almacen:
            if getattr(inv, 'fecha_vencimiento', None): 
                caducidades.append(inv.fecha_vencimiento)
        for inv in med.inventario_farmacia:
            if getattr(inv, 'fecha_vencimiento', None): 
                caducidades.append(inv.fecha_vencimiento)
        
        # Obtenemos la fecha más próxima a vencer y la formateamos de manera limpia
        if caducidades:
            proxima_caducidad = min(caducidades)
            caducidad_str = proxima_caducidad.strftime('%Y-%m-%d') if hasattr(proxima_caducidad, 'strftime') else str(proxima_caducidad)
        else:
            caducidad_str = "N/A"
            
        # 4. Lógica del semáforo visual según las existencias totales contra el mínimo
        if stock_total == 0:
            color_semaforo = "table-danger text-danger"
        elif stock_total <= (med.stock_minimo or 10):
            color_semaforo = "table-warning text-warning-emphasis"
        else:
            color_semaforo = "table-success text-success"

        medicamentos_con_stock.append({
            'clave': med.clave,
            'principio_activo': med.principio_activo,
            'presentacion': med.presentacion,
            'concentracion': med.concentracion,
            'lote': lotes_str,
            'caducidad': caducidad_str,
            'stock_almacen': stock_almacen,
            'stock_farmacia': stock_farmacia,
            'stock_total': stock_total,
            'color_semaforo': color_semaforo
        })
        
    return render_template('farmacia/detallegrupoterapeutico.html', grupo=grupo, medicamentos=medicamentos_con_stock)
    

@bp.route('/Medicamentos/TipoVia/<string:via_tipo>')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def ver_por_tipo_via(via_tipo):
    etiquetas = {
        'solidos': 'Medicamentos Sólidos y Formas Sólidas (Tabletas, Cápsulas, Grageas, Óvulos, etc.)',
        'liquidos': 'Soluciones, Líquidos y Suspensiones (Jarabes, Gotas, Suspensiones)',
        'inyectables': 'Inyectables y Vías Intravenosas (Ampolletas, Frascos Ámpula)'
    }
    
    query_meds = db.session.query(Medicamento)
    
    # RESTRICCIÓN INSTITUCIONAL: Solo jala productos cuyas claves inicien con 010 o 040
    query_meds = query_meds.filter(
        Medicamento.clave.like('010.%') | Medicamento.clave.like('040.%')
    )
    
    # CLASIFICACIÓN ESTRICTA CON EXCLUSIONES POR FORMA FARMACÉUTICA
    if via_tipo == 'solidos':
        query_meds = query_meds.filter(
            (
                Medicamento.via_administracion.ilike('%oral%') |
                Medicamento.via_administracion.ilike('%tableta%') |
                Medicamento.via_administracion.ilike('%capsula%') |
                Medicamento.via_administracion.ilike('%comprimido%') |
                Medicamento.via_administracion.ilike('%gragea%') |
                Medicamento.via_administracion.ilike('%ovulo%') |
                Medicamento.via_administracion.ilike('%polvo%') |
                Medicamento.via_administracion.ilike('%supositorio%') |
                Medicamento.presentacion.ilike('%tableta%') |
                Medicamento.presentacion.ilike('%capsula%') |
                Medicamento.presentacion.ilike('%comprimido%') |
                Medicamento.presentacion.ilike('%gragea%') |
                Medicamento.presentacion.ilike('%ovulo%')
            ) & 
            ~Medicamento.via_administracion.ilike('%solución%') &
            ~Medicamento.via_administracion.ilike('%suspensión%') &
            ~Medicamento.via_administracion.ilike('%jarabe%') &
            ~Medicamento.via_administracion.ilike('%gotas%') &
            ~Medicamento.via_administracion.ilike('%inyectable%') &
            ~Medicamento.presentacion.ilike('%solución%') &
            ~Medicamento.presentacion.ilike('%suspensión%') &
            ~Medicamento.presentacion.ilike('%jarabe%') &
            ~Medicamento.presentacion.ilike('%gotas%') &
            ~Medicamento.presentacion.ilike('%inyectable%')
        )

    elif via_tipo == 'liquidos':
        query_meds = query_meds.filter(
            (
                Medicamento.via_administracion.ilike('%solución%') |
                Medicamento.via_administracion.ilike('%líquido%') |
                Medicamento.via_administracion.ilike('%jarabe%') |
                Medicamento.via_administracion.ilike('%suspensión%') |
                Medicamento.via_administracion.ilike('%gotas%') |
                Medicamento.via_administracion.ilike('%elíxir%') |
                Medicamento.via_administracion.ilike('%emulsión%') |
                Medicamento.via_administracion.ilike('%spray%') |
                Medicamento.presentacion.ilike('%solución%') |
                Medicamento.presentacion.ilike('%jarabe%') |
                Medicamento.presentacion.ilike('%suspensión%') |
                Medicamento.presentacion.ilike('%gotas%') |
                Medicamento.presentacion.ilike('%líquido%')
            ) &
            ~Medicamento.via_administracion.ilike('%inyectable%') &
            ~Medicamento.via_administracion.ilike('%intravenosa%') &
            ~Medicamento.via_administracion.ilike('%intramuscular%') &
            ~Medicamento.presentacion.ilike('%ampolleta%') &
            ~Medicamento.presentacion.ilike('%frasco ámpula%')
        )

    elif via_tipo == 'inyectables':
        query_meds = query_meds.filter(
            Medicamento.via_administracion.ilike('%inyectable%') |
            Medicamento.via_administracion.ilike('%intravenosa%') |
            Medicamento.via_administracion.ilike('%ampolleta%') |
            Medicamento.via_administracion.ilike('%intramuscular%') |
            Medicamento.via_administracion.ilike('%subcutánea%') |
            Medicamento.presentacion.ilike('%ampolleta%') |
            Medicamento.presentacion.ilike('%frasco ámpula%') |
            Medicamento.presentacion.ilike('%inyectable%')
        )
    
    medicamentos_filtrados = query_meds.order_by(Medicamento.clave.asc()).all()
    
    # Mapeo de inventarios, lotes y semáforo visual (Alineación exacta con el HTML)
    datos_medicamentos = []
    for med in medicamentos_filtrados:
        stock_almacen = sum(inv.cantidad for inv in med.inventario_almacen if inv.cantidad)
        stock_farmacia = sum(inv.cantidad for inv in med.inventario_farmacia if inv.cantidad)
        stock_total = stock_almacen + stock_farmacia
        
        lotes = set()
        for inv in med.inventario_almacen:
            if getattr(inv, 'lote', None): lotes.add(inv.lote)
        for inv in med.inventario_farmacia:
            if getattr(inv, 'lote', None): lotes.add(inv.lote)
        lotes_str = ", ".join(lotes) if lotes else "N/A"
        
        caducidades = []
        for inv in med.inventario_almacen:
            if getattr(inv, 'fecha_vencimiento', None): caducidades.append(inv.fecha_vencimiento)
        for inv in med.inventario_farmacia:
            if getattr(inv, 'fecha_vencimiento', None): caducidades.append(inv.fecha_vencimiento)
            
        if caducidades:
            proxima_caducidad = min(caducidades)
            caducidad_str = proxima_caducidad.strftime('%Y-%m-%d') if hasattr(proxima_caducidad, 'strftime') else str(proxima_caducidad)
        else:
            caducidad_str = "N/A"
            
        if stock_total == 0:
            color_semaforo = "table-danger text-danger"
        elif stock_total <= (med.stock_minimo or 10):
            color_semaforo = "table-warning text-warning-emphasis"
        else:
            color_semaforo = "table-success text-success"

        datos_medicamentos.append({
            'clave': med.clave,
            'principio_activo': med.principio_activo,
            'presentacion': med.presentacion,
            'concentracion': med.concentracion,
            'lote': lotes_str,
            'caducidad': caducidad_str,
            'stock_almacen': stock_almacen,
            'stock_farmacia': stock_farmacia,
            'stock_total': stock_total,
            'color_semaforo': color_semaforo
        })

    titulo_vista = etiquetas.get(via_tipo, 'Listado de Insumos')
    return render_template('farmacia/tipovia.html', medicamentos=datos_medicamentos, titulo=titulo_vista)

    


    

#__________________________________________________________________________________________________________________________________

ancho_hoja, alto_hoja = letter

# Clase auxiliar para agregar numeración dinámica y contadores en el pie de página
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []
        # Variables para almacenar los totales
        self.total_meds = 0
        self.total_mats = 0

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 7.5)
        self.setFillColor(colors.HexColor("#4A5568"))
        
        # Formatear el texto de los contadores encontrados
        texto_contadores = f"{self.total_meds} Medicamentos | {self.total_mats} Materiales de Curación"
        fecha_str = datetime.utcnow().strftime("%d/%m/%Y %H:%M")
        
        # Pie de página alineado
        self.drawString(46, 20, texto_contadores)
        self.drawCentredString(ancho_hoja / 2, 20, f"Generado: {fecha_str} UTC")
        self.drawRightString(ancho_hoja - 46, 20, f"Página {self._pageNumber} de {page_count}")
        self.restoreState()


@bp.route('/inventario/catalogo')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def catalogo_medicamentos():
    # 1. Capturar el parámetro de búsqueda 'q' enviado desde la interfaz web
    query = request.args.get('q', '').strip()
    
    # Base de la consulta (Catálogo maestro sin paginación)
    medicamentos_query = Medicamento.query

    # Aplicar filtro parcial si el usuario ingresó un criterio de búsqueda
    if query:
        medicamentos_query = medicamentos_query.filter(
            (Medicamento.clave.ilike(f'%{query}%')) |
            (Medicamento.principio_activo.ilike(f'%{query}%'))
        )
    
    # Ejecutar la consulta en la base de datos
    medicamentos = medicamentos_query.all()

    catalogo_medicamentos = []
    catalogo_materiales = []

    # 2. Clasificación por prefijo de Clave
    for med in medicamentos:
        clave_limpia = (med.clave or "").strip()
        
        es_medicamento = clave_limpia.startswith("010") or clave_limpia.startswith("040")
        es_material = clave_limpia.startswith("060") or clave_limpia.startswith("080")

        if not (es_medicamento or es_material):
            continue

        item = {
            "clave": med.clave or "N/A",
            "nombre": med.principio_activo or "Sin Nombre",
            "presentacion": med.presentacion or "N/A",
            "concentracion": med.concentracion or "N/A"
        }

        if es_medicamento:
            catalogo_medicamentos.append(item)
        elif es_material:
            catalogo_materiales.append(item)

    # 3. Ordenamiento Alfabético (A-Z) estricto por Principio Activo / Nombre
    catalogo_medicamentos.sort(key=lambda x: x["nombre"].lower())
    catalogo_materiales.sort(key=lambda x: x["nombre"].lower())

    # 4. Configuración inicial del Documento PDF en memoria
    pdf_buffer = io.BytesIO()
    
    # Optimizamos los márgenes verticales (25 arriba y 35 abajo) para ganar espacio de impresión
    doc = SimpleDocTemplate(
        pdf_buffer, 
        pagesize=letter, 
        rightMargin=46, 
        leftMargin=46, 
        topMargin=25, 
        bottomMargin=35
    )
    story = []
    
    # Estilos de tipografía y colores corporativos
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle('T1', parent=styles['Heading1'], fontSize=16, leading=18, textColor=colors.HexColor('#1A365D'), spaceAfter=2)
    subtitulo_style = ParagraphStyle('T2', parent=styles['Normal'], fontSize=8.5, leading=11, textColor=colors.HexColor('#4A5568'), spaceAfter=8)
    seccion_style = ParagraphStyle('T3', parent=styles['Heading2'], fontSize=11, leading=14, textColor=colors.HexColor('#2B6CB0'), spaceBefore=8, spaceAfter=4, keepWithNext=True)
    
    # OPTIMIZACIÓN DE INTERLINEADO: Letra 7.5 y leading de 8.5 para máxima compresión horizontal y vertical
    th_style = ParagraphStyle('TH', parent=styles['Normal'], fontSize=8, leading=10, fontName='Helvetica-Bold', textColor=colors.white)
    td_style = ParagraphStyle('TD', parent=styles['Normal'], fontSize=7.5, leading=8.5, textColor=colors.HexColor('#2D3748'))

    # Encabezados del Reporte PDF
    story.append(Paragraph("Catálogo Institucional de Insumos Médicos", titulo_style))
    
    texto_subtitulo = "Listado completo de claves vigentes en el sistema."
    if query:
        texto_subtitulo = f"Resultados filtrados bajo el criterio: <b>'{query}'</b>."
    story.append(Paragraph(texto_subtitulo, subtitulo_style))

    # Definición estética de las tablas (Reducimos el padding a 2 para compactar filas)
    estilo_tabla = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2B6CB0')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E0')),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ])

    # Función interna para estructurar y dibujar los bloques de tablas continuas
    def agregar_bloque_catalogo(titulo_bloque, lista_elementos):
        story.append(Paragraph(titulo_bloque, seccion_style))
        if not lista_elementos:
            story.append(Paragraph("No se encontraron registros en esta categoría.", td_style))
            return

        # Estructura e inicialización de cabeceras
        filas_tabla = [[
            Paragraph("Clave", th_style),
            Paragraph("Principio Activo / Nombre", th_style),
            Paragraph("Presentación", th_style),
            Paragraph("Concentración", th_style)
        ]]

        # Inyección de registros de la base de datos
        for el in lista_elementos:
            filas_tabla.append([
                Paragraph(el["clave"], td_style),
                Paragraph(el["nombre"], td_style),
                Paragraph(el["presentacion"], td_style),
                Paragraph(el["concentracion"], td_style)
            ])

        # SOLUCCIÓN SYNTAX ERROR: Definición exacta de las columnas (Suman 520)
        anchos_columnas = [95, 185, 140, 100]
        
        # Constructor de tabla con división de filas habilitada contra LayoutError
        tabla_pdf = Table(
            filas_tabla, 
            colWidths=anchos_columnas, 
            repeatRows=1, 
            splitByRow=1
        )
        tabla_pdf.setStyle(estilo_tabla)
        story.append(tabla_pdf)

    # 5. Renderizar las secciones jerárquicas ordenadas
    agregar_bloque_catalogo("1. Medicamentos (Claves 010 / 040)", catalogo_medicamentos)
    story.append(Spacer(1, 5))  # Espaciador mínimo
    agregar_bloque_catalogo("2. Material de Curación (Claves 060 / 080)", catalogo_materiales)

    # 6. ENLAZAR RESULTADOS AL CANVAS
    def inicializar_canvas(*args, **kwargs):
        canvas_obj = NumberedCanvas(*args, **kwargs)
        canvas_obj.total_meds = len(catalogo_medicamentos)
        canvas_obj.total_mats = len(catalogo_materiales)
        return canvas_obj

    # Compilar y construir el flujo del documento usando el inicializador personalizado
    doc.build(story, canvasmaker=inicializar_canvas)

    # 7. Retornar el flujo de bytes binarios directo al navegador web
    pdf_buffer.seek(0)
    response = make_response(pdf_buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=catalogo_medicamentos.pdf'
    
    return response

@bp.route('/inventario/existencias')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def existencias_medicamentos():
    # 1. Capturar el parámetro de búsqueda 'q' enviado desde la interfaz web
    query = request.args.get('q', '').strip()
    
    # Cargar medicamentos e inventarios optimizando las relaciones
    medicamentos_query = Medicamento.query.options(
        selectinload(Medicamento.inventario_almacen),
        selectinload(Medicamento.inventario_farmacia)
    )

    # Aplicar filtro parcial si el usuario ingresó un criterio de búsqueda
    if query:
        medicamentos_query = medicamentos_query.filter(
            (Medicamento.clave.ilike(f'%{query}%')) |
            (Medicamento.principio_activo.ilike(f'%{query}%'))
        )
    
    medicamentos = medicamentos_query.all()

    reporte_medicamentos = []
    reporte_materiales = []

    # 2. Clasificación y cálculo de existencias totales por área
    for med in medicamentos:
        clave_limpia = (med.clave or "").strip()
        
        es_medicamento = clave_limpia.startswith("010") or clave_limpia.startswith("040")
        es_material = clave_limpia.startswith("060") or clave_limpia.startswith("080")

        if not (es_medicamento or es_material):
            continue

        # Sumar cantidades disponibles en Almacén (filtrando solo registros > 0)
        cant_alm = sum(i.cantidad for i in med.inventario_almacen if i.cantidad > 0)
        
        # Sumar cantidades disponibles en Farmacia (filtrando solo registros > 0)
        cant_far = sum(i.cantidad for i in med.inventario_farmacia if i.cantidad > 0)

        item = {
            "clave": med.clave or "N/A",
            "nombre": med.principio_activo or "Sin Nombre",
            "presentacion": med.presentacion or "N/A",
            "concentracion": med.concentracion or "N/A",
            "almacen": cant_alm,
            "farmacia": cant_far
        }

        if es_medicamento:
            reporte_medicamentos.append(item)
        elif es_material:
            reporte_materiales.append(item)

    # 3. Ordenamiento Alfabético (A-Z) estricto por Principio Activo
    reporte_medicamentos.sort(key=lambda x: x["nombre"].lower())
    reporte_materiales.sort(key=lambda x: x["nombre"].lower())

    # 4. Configuración del Documento PDF en memoria
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer, 
        pagesize=letter, 
        rightMargin=46, 
        leftMargin=46, 
        topMargin=25, 
        bottomMargin=35
    )
    story = []
    
    # Reutilizamos tus estilos ultra-compactos y optimizados de interlineado
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle('E_T1', parent=styles['Heading1'], fontSize=16, leading=18, textColor=colors.HexColor('#1A365D'), spaceAfter=2)
    subtitulo_style = ParagraphStyle('E_T2', parent=styles['Normal'], fontSize=8.5, leading=11, textColor=colors.HexColor('#4A5568'), spaceAfter=8)
    seccion_style = ParagraphStyle('E_T3', parent=styles['Heading2'], fontSize=11, leading=14, textColor=colors.HexColor('#2B6CB0'), spaceBefore=8, spaceAfter=4, keepWithNext=True)
    
    th_style = ParagraphStyle('E_TH', parent=styles['Normal'], fontSize=8, leading=10, fontName='Helvetica-Bold', textColor=colors.white)
    td_style = ParagraphStyle('E_TD', parent=styles['Normal'], fontSize=7.5, leading=8.5, textColor=colors.HexColor('#2D3748'))
    td_num = ParagraphStyle('E_TDN', parent=styles['Normal'], fontSize=7.5, leading=8.5, alignment=1, textColor=colors.HexColor('#2D3748')) # Centrado para números

    # Encabezados del Reporte
    story.append(Paragraph("Reporte de Existencias en Inventario", titulo_style))
    
    texto_subtitulo = "Saldos actuales consolidados por áreas de resguardo."
    if query:
        texto_subtitulo = f"Resultados filtrados bajo el criterio: <b>'{query}'</b>."
    story.append(Paragraph(texto_subtitulo, subtitulo_style))

    # Diseño estético de la tabla
    estilo_tabla = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2B6CB0')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E0')),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ])

    def agregar_bloque_existencias(titulo_bloque, lista_elementos):
        story.append(Paragraph(titulo_bloque, seccion_style))
        if not lista_elementos:
            story.append(Paragraph("No se encontraron registros en esta categoría.", td_style))
            return

        # Estructura de cabeceras incluyendo las dos nuevas columnas solicitadas
        filas_tabla = [[
            Paragraph("Clave", th_style),
            Paragraph("Principio Activo / Nombre", th_style),
            Paragraph("Presentación", th_style),
            Paragraph("Concentración", th_style),
            Paragraph("Almacén", th_style),
            Paragraph("Farmacia", th_style)
        ]]

        for el in lista_elementos:
            filas_tabla.append([
                Paragraph(el["clave"], td_style),
                Paragraph(el["nombre"], td_style),
                Paragraph(el["presentacion"], td_style),
                Paragraph(el["concentracion"], td_style),
                Paragraph(str(el["almacen"]), td_num), # Nueva columna Almacén
                Paragraph(str(el["farmacia"]), td_num)  # Nueva columna Farmacia
            ])

        # Redistribución exacta de anchos para meter 6 columnas en el límite de 520 puntos
        # Clave(85) + Nombre(185) + Presentación(100) + Concentración(70) + Almacén(40) + Farmacia(40) = 520 puntos
        anchos_columnas = [85, 185, 100, 70, 40, 40]
        
        tabla_pdf = Table(
            filas_tabla, 
            colWidths=anchos_columnas, 
            repeatRows=1, 
            splitByRow=1
        )
        tabla_pdf.setStyle(estilo_tabla)
        story.append(tabla_pdf)

    # 5. Renderizar bloques organizados
    agregar_bloque_existencias("1. Medicamentos (Claves 010 / 040)", reporte_medicamentos)
    story.append(Spacer(1, 5))
    agregar_bloque_existencias("2. Material de Curación (Claves 060 / 080)", reporte_materiales)

    # 6. Configurar e inicializar Canvas reutilizando el NumberedCanvas existente
    def inicializar_canvas(*args, **kwargs):
        canvas_obj = NumberedCanvas(*args, **kwargs)
        canvas_obj.total_meds = len(reporte_medicamentos)
        canvas_obj.total_mats = len(reporte_materiales)
        return canvas_obj

    doc.build(story, canvasmaker=inicializar_canvas)

    # 7. Respuesta binaria inline para el navegador
    pdf_buffer.seek(0)
    response = make_response(pdf_buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=existencias_inventario.pdf'
    
    return response


#===============================================================================ENTRADA AL ALMACEN================================

@bp.route('/entradas/listar')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def listar_entradas():
    query = request.args.get('q', '').strip().upper()
    
    # Base de la consulta con JOIN para poder buscar por Clave o Principio Activo
    entradas_query = EntradaAlmacen.query.join(Medicamento)

    if query:
        entradas_query = entradas_query.filter(
            (Medicamento.clave.ilike(f'%{query}%')) |
            (Medicamento.principio_activo.ilike(f'%{query}%')) |
            (EntradaAlmacen.lote.ilike(f'%{query}%')) |
            (EntradaAlmacen.proveedor.ilike(f'%{query}%'))
        )

    # Ordenar por las más recientes primero
    entradas = entradas_query.order_by(EntradaAlmacen.fecha_entrada.desc()).limit(10).all()

    
    return render_template('farmacia/entradas.html', 
                           entradas=entradas, 
                           query=query)



@bp.route('/descargar_oc99')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def descargar_oc99():
    try:
        # 1. CAPTURAR DATOS DEL FILTRO (Vía request.args desde JS/SweetAlert2)
        ahora = datetime.now()
        anio = request.args.get('anio', ahora.year, type=int)
        mes = request.args.get('mes', ahora.month, type=int)

        # 2. CALCULAR DÍAS DEL MES SELECCIONADO
        dias_mes = calendar.monthrange(anio, mes)[1]

        # 3. OBTENER MEDICAMENTOS ÚNICOS
        medicamentos = db.session.query(Medicamento).order_by(Medicamento.clave).all()

        # 4. OBTENER ENTRADAS FILTRADAS POR MES Y AÑO
        entradas = db.session.query(EntradaAlmacen).join(Medicamento).filter(
            db.extract('year', EntradaAlmacen.fecha_entrada) == anio,
            db.extract('month', EntradaAlmacen.fecha_entrada) == mes
        ).all()

        # 5. CREAR MATRIZ DE DATOS
        matriz = {}
        for med in medicamentos:
            # CONCATENAMOS: "PRINCIPIO ACTIVO - PRESENTACIÓN"
            # .strip() elimina espacios accidentales y .upper() estandariza
            nombre_full = f"{med.principio_activo or ''} - {med.presentacion or ''}".strip().upper()
            
            matriz[med.id_medicamento] = {
                'clave': med.clave,
                'nombre_completo': nombre_full,
                'dias': {d: "" for d in range(1, dias_mes + 1)}  # Inicializado en vacío
            }

        # Llenar la matriz con las cantidades de las entradas
        for e in entradas:
            dia = e.fecha_entrada.day
            if e.id_medicamento in matriz:
                # Recuperar valor previo controlando si está vacío
                valor_actual = matriz[e.id_medicamento]['dias'][dia]
                cantidad_previa = valor_actual if isinstance(valor_actual, int) else 0
                
                matriz[e.id_medicamento]['dias'][dia] = cantidad_previa + e.cantidad

        # 6. CONSTRUIR EL ARCHIVO EXCEL
        wb = Workbook()
        ws = wb.active
        ws.title = f"OC99_{mes}_{anio}"

        # ENCABEZADOS
        headers = ["CLAVE", "MEDICAMENTO (PRINCIPIO - PRESENTACIÓN)"] + [str(d) for d in range(1, dias_mes + 1)]
        ws.append(headers)

        # Estilo para los encabezados (Verde, Negrita, Texto Blanco)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        # LLENADO DE FILAS
        for med in matriz.values():
            fila = [
                med['clave'],
                med['nombre_completo']
            ] + [med['dias'][d] for d in range(1, dias_mes + 1)]
            ws.append(fila)

        # 7. AJUSTES FINALES DE DISEÑO
        ws.column_dimensions['A'].width = 15  # Clave
        ws.column_dimensions['B'].width = 70  # Nombre concatenado (más ancho)
        
        # Ajustar ancho de las columnas de los días (columnas C en adelante)
        for col_idx in range(3, 3 + dias_mes):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 4
            # Alinear números de los días al centro
            for cell in ws[col_letter]:
                cell.alignment = center_alignment

        # 8. PREPARAR DESCARGA
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        nombre_archivo = f"REPORTE_OC99_{mes:02d}_{anio}.xlsx"

        return send_file(
            output,
            download_name=nombre_archivo,
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        flash(f"❌ Error al generar el reporte: {str(e)}", "danger")
        return redirect(url_for('farmacia.reporte_inventario'))

@bp.route('/descargar_entradas')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def descargar_entradas():
    # 1. Capturar parámetros de fecha (vienen como 'YYYY-MM-DD')
    fecha_inicio_str = request.args.get('inicio')
    fecha_fin_str = request.args.get('fin')

    # Convertir strings a objetos date de Python
    from datetime import datetime
    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        # Fallback por si algo falla: mes actual
        hoy = datetime.now()
        fecha_inicio = hoy.replace(day=1).date()
        import calendar
        ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
        fecha_fin = hoy.replace(day=ultimo_dia).date()

    # 2. Consulta a la base de datos usando el rango (between)
    entradas = db.session.query(
        Medicamento.clave,
        Medicamento.principio_activo,
        Medicamento.presentacion,
        EntradaAlmacen.lote,
        EntradaAlmacen.cantidad,
        EntradaAlmacen.fecha_caducidad,
        EntradaAlmacen.fecha_entrada # Necesario para el nombre del reporte si quieres
    ).join(Medicamento).filter(
        EntradaAlmacen.fecha_entrada.between(fecha_inicio, fecha_fin)
    ).all()

    # 3. Crear el Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Entradas"

    # Encabezados
    headers = ["CLAVE", "MEDICAMENTO (PRINCIPIO - PRESENTACIÓN)", "LOTE", "CANTIDAD", "FECHA DE CADUCIDAD"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 4. Llenar los datos
    for e in entradas:
        nombre_completo = f"{e.principio_activo or ''} - {e.presentacion or ''}".strip().upper()
        
        ws.append([
            e.clave,
            nombre_completo,
            e.lote,
            e.cantidad,
            e.fecha_caducidad.strftime('%d/%m/%Y') if e.fecha_caducidad else "N/A"
        ])

    # 5. Ajustar anchos
    column_widths = {'A': 15, 'B': 60, 'C': 15, 'D': 12, 'E': 20}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 6. Preparar descarga
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nombre dinámico con el rango de fechas
    nombre_archivo = f"ENTRADAS_{fecha_inicio.strftime('%d%m%Y')}_A_{fecha_fin.strftime('%d%m%Y')}.xlsx"

    return send_file(
        output,
        download_name=nombre_archivo,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@bp.route('/entradas/nueva', methods=['GET', 'POST'])
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def nueva_entrada():
    if request.method == 'GET':
        return render_template('farmacia/nueva_entrada.html')

    # Inicializamos las variables fuera del try para que estén disponibles en el except
    destino = request.form.get('destino')
    proveedor = request.form.get('proveedor', '').strip().upper()
    obs_general = request.form.get('observaciones_general', '').strip().upper()

    ids_medicamentos = request.form.getlist('id_medicamento[]')
    lotes = request.form.getlist('lote[]')
    cantidades = request.form.getlist('cantidad[]')
    fechas_cad = request.form.getlist('fecha_caducidad[]')
    notas_items = request.form.getlist('notas_item[]')

    try:
        if not ids_medicamentos:
            flash("❌ No hay datos para registrar.", "warning")
            # Levantamos un error controlado para reutilizar la lógica de retorno del except
            raise ValueError("La lista de medicamentos está vacía.")

        # 3. Procesamiento en Bloque
        for i in range(len(ids_medicamentos)):
            id_med = int(ids_medicamentos[i])
            cant = int(cantidades[i])
            lote_actual = lotes[i].strip().upper()
            f_cad = datetime.strptime(fechas_cad[i], '%Y-%m-%d').date()
            nota_actual = f"{obs_general} | {notas_items[i]}".strip(" | ").upper()

            # A. Historial Global
            nueva_entrada_log = EntradaAlmacen(
                id_medicamento=id_med,
                cantidad=cant,
                lote=lote_actual,
                fecha_caducidad=f_cad,
                fecha_entrada=datetime.utcnow(),
                proveedor=proveedor,
                observaciones=f"ENTRADA DIRECTA A {destino.upper()} - {nota_actual}",
                id_usuario=current_user.id_usuario
            )
            db.session.add(nueva_entrada_log)

            # B. Actualización de Stock
            modelo_inventario = InventarioFarmacia if destino == 'farmacia' else InventarioAlmacen

            stock_item = modelo_inventario.query.filter_by(
                id_medicamento=id_med, 
                lote=lote_actual
            ).first()

            if stock_item:
                stock_item.cantidad += cant
            else:
                nuevo_stock = modelo_inventario(
                    id_medicamento=id_med,
                    cantidad=cant,
                    lote=lote_actual,
                    fecha_vencimiento=f_cad
                )
                db.session.add(nuevo_stock)

        # 4. Confirmación única de la transacción
        db.session.commit()
        flash(f"✅ Éxito: {len(ids_medicamentos)} registros guardados en {destino.upper()}.", "success")
        return redirect(url_for('farmacia.listar_entradas'))

    except Exception as e:
        db.session.rollback()
        
        # Evitamos duplicar el flash si ya pusimos el de lista vacía
        if "La lista de medicamentos está vacía." not in str(e):
            flash(f"❌ Error en el proceso masivo: {str(e)}", "danger")
        
        # Recuperamos los nombres reales de la BD para reconstruir el Select2
        textos_medicamentos = []
        for id_med in ids_medicamentos:
            if id_med.isdigit():
                # Reemplaza 'Medicamento' por tu modelo real y 'nombre' por tu columna de texto/clave
                med = Medicamento.query.get(int(id_med))
                textos_medicamentos.append(f"{med.clave} - {med.principio_activo}" if med else f"ID: {id_med}")
            else:
                textos_medicamentos.append("Seleccionar Medicamento")

        # Empaquetamos los datos de las filas
        datos_previos = {
            'id_medicamento': ids_medicamentos,
            'texto_medicamento': textos_medicamentos,
            'lote': lotes,
            'cantidad': cantidades,
            'fecha_caducidad': fechas_cad,
            'notas_item': notas_items
        }

        # Retornamos la plantilla inyectando lo capturado (Cabecera y Tabla)
        return render_template(
            'farmacia/nueva_entrada.html',
            destino_previo=destino,
            proveedor_previo=proveedor,
            observaciones_previo=obs_general,
            datos_previos=datos_previos
        )

@bp.route('/entradas/editar/<int:id_entrada>', methods=['GET', 'POST'])
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def editar_entrada(id_entrada):
    entrada = EntradaAlmacen.query.get_or_404(id_entrada)
    medicamentos = Medicamento.query.order_by(Medicamento.principio_activo).all()

    if request.method == 'POST':
        try:
            # 1. Guardar valores antiguos para calcular la diferencia
            lote_anterior = entrada.lote
            cantidad_anterior = entrada.cantidad
            id_med_anterior = entrada.id_medicamento

            # 2. Capturar nuevos valores
            nueva_cantidad = int(request.form['cantidad'])
            nuevo_lote = request.form.get('lote', '').strip().upper()
            f_caducidad_raw = request.form.get('fecha_caducidad')
            nueva_fecha_caducidad = datetime.strptime(f_caducidad_raw, '%Y-%m-%d').date() if f_caducidad_raw else None

            # 3. REVERTIR STOCK ANTERIOR (Limpiamos el rastro de la entrada antes del cambio)
            inv_anterior = InventarioAlmacen.query.filter_by(
                id_medicamento=id_med_anterior, 
                lote=lote_anterior
            ).first()
            
            if inv_anterior:
                inv_anterior.cantidad -= cantidad_anterior

            # 4. ACTUALIZAR REGISTRO DE ENTRADA
            entrada.id_medicamento = int(request.form['id_medicamento'])
            entrada.cantidad = nueva_cantidad
            entrada.lote = nuevo_lote
            entrada.fecha_caducidad = nueva_fecha_caducidad
            entrada.proveedor = request.form.get('proveedor', '').upper()
            entrada.observaciones = request.form.get('observaciones', '').upper()

            # 5. APLICAR NUEVO STOCK
            inv_nuevo = InventarioAlmacen.query.filter_by(
                id_medicamento=entrada.id_medicamento, 
                lote=nuevo_lote
            ).first()

            if inv_nuevo:
                inv_nuevo.cantidad += nueva_cantidad
                inv_nuevo.fecha_vencimiento = nueva_fecha_caducidad # Actualizamos caducidad por si cambió
            else:
                inv_nuevo = InventarioAlmacen(
                    id_medicamento=entrada.id_medicamento,
                    cantidad=nueva_cantidad,
                    lote=nuevo_lote,
                    fecha_vencimiento=nueva_fecha_caducidad
                )
                db.session.add(inv_nuevo)

            db.session.commit()
            flash("✅ Entrada e Inventario actualizados correctamente.", "success")
            return redirect(url_for('farmacia.listar_entradas'))

        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error al modificar: {str(e)}", "danger")

    return render_template('farmacia/editar_entrada.html', entrada=entrada, medicamentos=medicamentos)


@bp.route('/entradas/eliminar/<int:id_entrada>', methods=['POST'])
@roles_required(['Administrador'])
def eliminar_entrada(id_entrada):
    entrada = EntradaAlmacen.query.get_or_404(id_entrada)
    
    # 3. REVERTIR STOCK (Solo si hay existencia suficiente)
    inv = InventarioAlmacen.query.filter_by(id_medicamento=entrada.id_medicamento, lote=entrada.lote).first()
    
    if inv and inv.cantidad >= entrada.cantidad:
        inv.cantidad -= entrada.cantidad
        db.session.delete(entrada)
        db.session.commit()
        flash("✅ Entrada eliminada y stock revertido.", "success")
    else:
        flash("❌ No se puede eliminar: El stock de este lote ya es menor a la cantidad de esta entrada.", "danger")
        
    return redirect(url_for('farmacia.listar_entradas'))

#))))))))))))))))))))))))))))))))))))))))))))))))))))))))))))))))) BUSQUEDAS  ))))))))))))))))))))))))))))))))))))))))))))))

@bp.route('/medicamentos/buscar')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def buscar_medicamentos():
    q = request.args.get('q', '').strip().upper() # Buscamos en mayúsculas para coincidir
    if not q:
        return jsonify([])

    # Filtramos solo por los campos que SI existen en tu modelo
    resultados = Medicamento.query.filter(
        (Medicamento.clave.ilike(f"%{q}%")) |
        (Medicamento.principio_activo.ilike(f"%{q}%"))
    ).limit(20).all()

    # Formateamos para que Select2 muestre la info clara
    data = []
    for med in resultados:
        data.append({
            'id': med.id_medicamento,
            # Mostramos Clave y Principio Activo únicamente
            'text': f"{med.clave} | {med.principio_activo} | {med.presentacion} | {med.concentracion}"
        })

    return jsonify(data)



@bp.route('/buscar_lotes_almacen/<int:id_medicamento>')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def buscar_lotes_almacen(id_medicamento):
    # Buscamos en InventarioAlmacen los registros con cantidad > 0
    lotes = InventarioAlmacen.query.filter(
        InventarioAlmacen.id_medicamento == id_medicamento,
        InventarioAlmacen.cantidad > 0
    ).all()
    
    return jsonify([{
        'lote': l.lote, 
        'cantidad': l.cantidad, 
        'vencimiento': l.fecha_vencimiento.strftime('%d/%m/%Y') if l.fecha_vencimiento else 'N/A'
    } for l in lotes])



#===============================================================================MOVIMIENTO DE ALMACEN A FARMACIA================================
@bp.route('/movimientos/almacen')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def listar_movimientos():
    ahora = datetime.now() # <--- DEFINIDO para evitar errores
    query = request.args.get('q', '').strip().upper()
    
    movimientos_query = MovimientoAlmacenFarmacia.query.join(Medicamento).join(Usuario)

    if query:
        movimientos_query = movimientos_query.filter(
            (Medicamento.clave.ilike(f'%{query}%')) |
            (Medicamento.principio_activo.ilike(f'%{query}%')) |
            (Usuario.usuario.ilike(f'%{query}%')) |
            (MovimientoAlmacenFarmacia.observaciones.ilike(f'%{query}%'))
        )

    # Nota: Removí el .limit(10) para que la tabla muestre los resultados completos o paginados si aplica
    movimientos = movimientos_query.order_by(MovimientoAlmacenFarmacia.fecha_movimiento.desc()).all()
        
    return render_template('farmacia/movimientos.html', 
                           movimientos=movimientos, 
                           query=query, 
                           mes_actual=ahora.month, 
                           anio_actual=ahora.year)
		

@bp.route('/movimientos/nuevo_movimiento', methods=['GET', 'POST'])
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def nuevo_movimiento():
    if request.method == 'GET':
        return render_template('farmacia/nuevo_movimiento.html')

    if request.method == 'POST':
        try:
            # 1. Obtener datos de cabecera
            tipo_movimiento = request.form.get('tipo_movimiento')
            destino_externo = request.form.get('destino_externo', '').strip().upper()
            obs_general = request.form.get('obs_general', '').strip().upper()
            
            ids_medicamentos = request.form.getlist('id_medicamento[]')
            lotes = request.form.getlist('lote[]')
            cantidades = request.form.getlist('cantidad[]')

            if not ids_medicamentos:
                flash("❌ No se seleccionaron productos.", "warning")
                return redirect(url_for('farmacia.nuevo_movimiento'))

            # Construir observación detallada
            obs_final = f"{tipo_movimiento}: {obs_general}"
            if destino_externo:
                obs_final += f" - DESTINO: {destino_externo}"

            # 2. Procesar cada fila
            for i in range(len(ids_medicamentos)):
                id_med = int(ids_medicamentos[i])
                lote_actual = lotes[i].strip().upper()
                cant_a_trasladar = int(cantidades[i])
                
                inv_almacen = InventarioAlmacen.query.filter_by(
                    id_medicamento=id_med, 
                    lote=lote_actual
                ).first()

                if not inv_almacen or inv_almacen.cantidad < cant_a_trasladar:
                    nombre_med = inv_almacen.medicamento.principio_activo if inv_almacen else f"ID {id_med}"
                    raise Exception(f"Stock insuficiente para {nombre_med} (Lote: {lote_actual}).")

                # A. Historial
                movimiento = MovimientoAlmacenFarmacia(
                    id_medicamento=id_med,
                    cantidad=cant_a_trasladar,
                    lote=lote_actual,
                    fecha_vencimiento=inv_almacen.fecha_vencimiento,
                    fecha_movimiento=datetime.utcnow(),
                    id_usuario=current_user.id_usuario,
                    observaciones=obs_final.strip()
                )
                db.session.add(movimiento)

                # B. Restar de Almacén
                inv_almacen.cantidad -= cant_a_trasladar

                # C. Sumar a Farmacia (Solo si es traslado interno)
                if tipo_movimiento == 'TRASLADO_FARMACIA':
                    inv_farmacia = InventarioFarmacia.query.filter_by(
                        id_medicamento=id_med, 
                        lote=lote_actual
                    ).first()

                    if inv_farmacia:
                        inv_farmacia.cantidad += cant_a_trasladar
                    else:
                        inv_farmacia = InventarioFarmacia(
                            id_medicamento=id_med,
                            cantidad=cant_a_trasladar,
                            lote=lote_actual,
                            fecha_vencimiento=inv_almacen.fecha_vencimiento
                        )
                        db.session.add(inv_farmacia)

            db.session.commit()
            flash(f"✅ Operación {tipo_movimiento} procesada correctamente.", "success")
            return redirect(url_for('farmacia.listar_movimientos'))

        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error: {str(e)}", "danger")
            return redirect(url_for('farmacia.nuevo_movimiento'))

    return render_template('farmacia/nuevo_movimiento.html')

@bp.route('/movimientos/editar/<int:id>', methods=['GET', 'POST'])
@roles_required(['Administrador'])
def editar_movimiento(id):
    movimiento = MovimientoAlmacenFarmacia.query.get_or_404(id)
    cantidad_anterior = movimiento.cantidad
    medicamentos = Medicamento.query.order_by(Medicamento.principio_activo).all()

    if request.method == 'POST':
        try:
            nueva_cantidad = int(request.form['cantidad'])
            diferencia = nueva_cantidad - cantidad_anterior
            
            # Detectar el tipo de movimiento desde las observaciones o un campo tipo si lo agregaste
            # Aquí usamos una lógica simple: si existe en farmacia y la obs dice TRASLADO, es un traslado.
            es_traslado = "TRASLADO_FARMACIA" in movimiento.observaciones

            # 1. Buscar registros de inventario
            inv_almacen = InventarioAlmacen.query.filter_by(id_medicamento=movimiento.id_medicamento, lote=movimiento.lote).first()
            inv_farmacia = InventarioFarmacia.query.filter_by(id_medicamento=movimiento.id_medicamento, lote=movimiento.lote).first()

            # --- VALIDACIONES ---
            
            # A. Si la cantidad AUMENTA: Validar que el Almacén tenga el extra
            if diferencia > 0:
                if not inv_almacen or inv_almacen.cantidad < diferencia:
                    stock_disp = inv_almacen.cantidad if inv_almacen else 0
                    raise Exception(f"Stock insuficiente en almacén. Disponible extra: {stock_disp}")

            # B. Si la cantidad DISMINUYE y es TRASLADO: Validar que Farmacia tenga para devolver
            if diferencia < 0 and es_traslado:
                if not inv_farmacia or inv_farmacia.cantidad < abs(diferencia):
                    raise Exception("No se puede reducir: Farmacia ya consumió/vendió parte de este lote.")

            # --- APLICAR CAMBIOS ---

            # 1. Afectar Almacén (Siempre se afecta)
            if inv_almacen:
                inv_almacen.cantidad -= diferencia
            else:
                # Caso extremo: el registro de almacén desapareció (se borró el lote)
                raise Exception("El registro de lote en almacén ya no existe. No se puede editar.")

            # 2. Afectar Farmacia (SOLO si el movimiento original fue un traslado)
            if es_traslado:
                if inv_farmacia:
                    inv_farmacia.cantidad += diferencia
                else:
                    # Si no existe en farmacia pero es traslado, lo creamos (caso raro)
                    inv_farmacia = InventarioFarmacia(
                        id_medicamento=movimiento.id_medicamento,
                        cantidad=nueva_cantidad,
                        lote=movimiento.lote,
                        fecha_vencimiento=movimiento.fecha_vencimiento
                    )
                    db.session.add(inv_farmacia)

            # 3. Actualizar el registro del movimiento
            movimiento.cantidad = nueva_cantidad
            movimiento.observaciones = request.form.get('observaciones', '').upper()
            
            db.session.commit()
            flash("✅ Movimiento y stocks actualizados correctamente.", "success")
            return redirect(url_for('farmacia.listar_movimientos'))

        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error: {str(e)}", "danger")
            return redirect(url_for('farmacia.editar_movimiento', id=id))

    return render_template('farmacia/editar_movimiento.html', movimiento=movimiento, medicamentos=medicamentos)

@bp.route('/movimientos/eliminar/<int:id>', methods=['POST'])
@roles_required(['Administrador'])
def eliminar_movimiento(id):
    movimiento = MovimientoAlmacenFarmacia.query.get_or_404(id)
    try:
        # 1. Identificar si el movimiento afectó a la farmacia
        # Basado en la etiqueta que pusimos en el proceso masivo
        es_traslado = "TRASLADO_FARMACIA" in (movimiento.observaciones or "")

        inv_almacen = InventarioAlmacen.query.filter_by(
            id_medicamento=movimiento.id_medicamento, 
            lote=movimiento.lote
        ).first()

        # 2. Lógica de reversión para Traslados a Farmacia
        if es_traslado:
            inv_farmacia = InventarioFarmacia.query.filter_by(
                id_medicamento=movimiento.id_medicamento, 
                lote=movimiento.lote
            ).first()

            # Validar si la farmacia tiene suficiente para devolver
            if not inv_farmacia or inv_farmacia.cantidad < movimiento.cantidad:
                raise Exception("No se puede eliminar: el stock ya fue consumido o no existe en farmacia.")
            
            # Restamos de farmacia
            inv_farmacia.cantidad -= movimiento.cantidad
        
        # 3. Devolver siempre al Almacén (independientemente del tipo de movimiento)
        if inv_almacen:
            inv_almacen.cantidad += movimiento.cantidad
        else:
            # Si por alguna razón el registro del lote desapareció de almacén, lo recreamos
            nuevo_inv_almacen = InventarioAlmacen(
                id_medicamento=movimiento.id_medicamento,
                cantidad=movimiento.cantidad,
                lote=movimiento.lote,
                fecha_vencimiento=movimiento.fecha_vencimiento
            )
            db.session.add(nuevo_inv_almacen)

        # 4. Borrar el registro del historial
        db.session.delete(movimiento)
        db.session.commit()
        
        mensaje = "Movimiento anulado. El stock ha retornado al Almacén"
        if es_traslado: mensaje += " y se descontó de Farmacia."
        else: mensaje += " (Baja/Transferencia cancelada)."
        
        flash(f"✅ {mensaje}", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al eliminar: {str(e)}", "danger")
    
    return redirect(url_for('farmacia.listar_movimientos'))



@bp.route('/descargar_traspasos_oc99')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def descargar_traspasos_oc99():
    from datetime import datetime
    import calendar
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import request, send_file
    
    ahora = datetime.now()
    anio = request.args.get('anio', ahora.year, type=int)
    mes = request.args.get('mes', ahora.month, type=int)
    dias_mes = calendar.monthrange(anio, mes)[1]

    # 1. Obtener datos
    medicamentos = db.session.query(Medicamento).order_by(Medicamento.clave).all()
    traspasos = db.session.query(MovimientoAlmacenFarmacia).filter(
        db.extract('year', MovimientoAlmacenFarmacia.fecha_movimiento) == anio,
        db.extract('month', MovimientoAlmacenFarmacia.fecha_movimiento) == mes
    ).all()

    # 2. Matriz de datos y Totales por día
    matriz = {}
    totales_piezas_dia = {d: 0 for d in range(1, dias_mes + 1)}
    claves_por_dia = {d: set() for d in range(1, dias_mes + 1)}

    for med in medicamentos:
        nombre_completo = f"{med.principio_activo or ''} - {med.presentacion or ''}".strip().upper()
        matriz[med.id_medicamento] = {
            'clave': med.clave,
            'nombre': nombre_completo,
            'dias': {d: "" for d in range(1, dias_mes + 1)}  # Inicializado en vacío
        }

    for t in traspasos:
        if t.fecha_movimiento:
            dia = t.fecha_movimiento.day
            if t.id_medicamento in matriz:
                # Recuperar valor previo controlando si está vacío
                valor_actual = matriz[t.id_medicamento]['dias'][dia]
                cantidad_previa = valor_actual if isinstance(valor_actual, int) else 0
                
                matriz[t.id_medicamento]['dias'][dia] = cantidad_previa + t.cantidad
                totales_piezas_dia[dia] += t.cantidad
                if t.cantidad > 0:
                    claves_por_dia[dia].add(t.id_medicamento)

    # 3. Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "OC99 Traspasos"

    # --- CUADRO DE RESUMEN SUPERIOR (Como en la imagen) ---
    font_bold = Font(bold=True)
    fill_resumen = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Fila de TOTAL CLAVES (Vacío si es 0)
    fila_claves = ["TOTAL CLAVES:", ""] + [len(claves_por_dia[d]) if len(claves_por_dia[d]) > 0 else "" for d in range(1, dias_mes + 1)]
    ws.append(fila_claves)
    
    # Fila de TOTAL PIEZAS (Vacío si es 0)
    fila_piezas = ["TOTAL PIEZAS:", ""] + [totales_piezas_dia[d] if totales_piezas_dia[d] > 0 else "" for d in range(1, dias_mes + 1)]
    ws.append(fila_piezas)

    # Estilo para el resumen
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = font_bold
            cell.fill = fill_resumen
            cell.alignment = Alignment(horizontal="center")

    ws.append([]) # Espacio en blanco

    # 4. ENCABEZADOS DE LA TABLA
    headers = ["CLAVE", "MEDICAMENTO"] + [str(d) for d in range(1, dias_mes + 1)]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid") # Naranja
    for cell in ws[4]: # La fila 4 es ahora el encabezado
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 5. DATOS DE MEDICAMENTOS
    for med in matriz.values():
        fila = [med['clave'], med['nombre']] + [med['dias'][d] for d in range(1, dias_mes + 1)]
        ws.append(fila)

    # 6. AJUSTES FINALES
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 60
    
    # Ajustar ancho de columnas de días
    for col_idx in range(3, 3 + dias_mes):
        ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = 5

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name=f"OC99_TRASPASOS_{mes:02d}_{anio}.xlsx", as_attachment=True)
    
    
    
from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from datetime import datetime
import calendar
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# --- RUTA 1: VISTA EN PÁGINA WEB (TABLA HTML) ---
@bp.route('/movimientos/ver_tabla')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def ver_movimientos_tabla():
    tipo = request.args.get('tipo', 'TODOS')
    
    query = db.session.query(MovimientoAlmacenFarmacia, Medicamento)\
        .join(Medicamento, MovimientoAlmacenFarmacia.id_medicamento == Medicamento.id_medicamento)
    
    # Si seleccionan CADUCADO, buscamos la etiqueta exacta de tu BD: BAJA_CADUCIDAD
    if tipo != 'TODOS':
        if tipo == 'CADUCADO':
            query = query.filter(MovimientoAlmacenFarmacia.observaciones.like("BAJA_CADUCIDAD%"))
        else:
            query = query.filter(MovimientoAlmacenFarmacia.observaciones.like(f"{tipo}%"))
        
    movimientos = query.order_by(MovimientoAlmacenFarmacia.fecha_movimiento.desc()).all()
    
    return render_template('farmacia/movimientos_tabla.html', movimientos=movimientos, tipo_actual=tipo)


# --- RUTA 2: DESCARGA DE EXCEL MATRICIAL O COMPLETO ---
@bp.route('/movimientos/descargar_excel')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def descargar_movimientos_excel():
    ahora = datetime.now()
    rango = request.args.get('rango', 'todos')
    tipo_movimiento = request.args.get('tipo_movimiento', 'TODOS')
    
    query_traspasos = db.session.query(MovimientoAlmacenFarmacia)
    
    # Ajuste de filtro para BAJA_CADUCIDAD
    if tipo_movimiento != 'TODOS':
        if tipo_movimiento == 'CADUCADO':
            query_traspasos = query_traspasos.filter(MovimientoAlmacenFarmacia.observaciones.like("BAJA_CADUCIDAD%"))
        else:
            query_traspasos = query_traspasos.filter(MovimientoAlmacenFarmacia.observaciones.like(f"{tipo_movimiento}%"))

    # REPORTE MATRICIAL POR DÍAS (MENSUAL)
    if rango == 'mes':
        anio = request.args.get('anio', ahora.year, type=int)
        mes = request.args.get('mes', ahora.month, type=int)
        dias_mes = calendar.monthrange(anio, mes)[1]

        traspasos = query_traspasos.filter(
            db.extract('year', MovimientoAlmacenFarmacia.fecha_movimiento) == anio,
            db.extract('month', MovimientoAlmacenFarmacia.fecha_movimiento) == mes
        ).all()

        medicamentos = db.session.query(Medicamento).order_by(Medicamento.clave).all()

        matriz = {}
        totales_piezas_dia = {d: 0 for d in range(1, dias_mes + 1)}
        claves_por_dia = {d: set() for d in range(1, dias_mes + 1)}

        for med in medicamentos:
            nombre_completo = f"{med.principio_activo or ''} - {med.presentacion or ''}".strip().upper()
            matriz[med.id_medicamento] = {
                'clave': med.clave,
                'nombre': nombre_completo,
                'dias': {d: 0 for d in range(1, dias_mes + 1)}
            }

        for t in traspasos:
            dia = t.fecha_movimiento.day
            if t.id_medicamento in matriz:
                matriz[t.id_medicamento]['dias'][dia] += t.cantidad
                totales_piezas_dia[dia] += t.cantidad
                if t.cantidad > 0:
                    claves_por_dia[dia].add(t.id_medicamento)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Reporte {mes}_{anio}"

        font_bold = Font(bold=True)
        fill_resumen = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        ws.append(["TOTAL CLAVES:", ""] + [len(claves_por_dia[d]) for d in range(1, dias_mes + 1)])
        ws.append(["TOTAL PIEZAS:", ""] + [totales_piezas_dia[d] for d in range(1, dias_mes + 1)])

        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.font = font_bold
                cell.fill = fill_resumen
                cell.alignment = Alignment(horizontal="center")

        ws.append([]) 

        headers = ["CLAVE", "MEDICAMENTO"] + [str(d) for d in range(1, dias_mes + 1)]
        ws.append(headers)
        
        color_header = "198754" 
        if tipo_movimiento == "CADUCADO": color_header = "DC3545" 
        if tipo_movimiento == "EXTRAVIO": color_header = "FFC107" 

        header_fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
        for cell in ws[4]:
            cell.font = Font(bold=True, color="FFFFFF" if color_header != "FFC107" else "000000")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for med in matriz.values():
            fila = [med['clave'], med['nombre']] + [med['dias'][d] for d in range(1, dias_mes + 1)]
            ws.append(fila)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 60
        for col_idx in range(3, 3 + dias_mes):
            ws.column_dimensions[ws.cell(row=4, column=col_idx).column_letter].width = 5

        filename = f"MATRIZ_{tipo_movimiento}_{mes:02d}_{anio}.xlsx"

    # REPORTE HISTÓRICO LINEAL COMPLETO (Si eligen Todo)
    else:
        traspasos = query_traspasos.order_by(MovimientoAlmacenFarmacia.fecha_movimiento.desc()).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial"
        
        headers = ["FECHA", "CLAVE", "MEDICAMENTO", "LOTE", "CANTIDAD", "OBSERVACIONES"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            
        for t in traspasos:
            med = t.medicamento 
            nombre = f"{med.principio_activo} - {med.presentacion}".upper() if med else "DESCONOCIDO"
            clave = med.clave if med else "N/A"
            ws.append([t.fecha_movimiento.strftime('%d/%m/%Y %H:%M'), clave, nombre, t.lote, t.cantidad, t.observaciones])
            
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 50
        
        filename = f"HISTORIAL_{tipo_movimiento}.xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=filename, as_attachment=True)

#=========================================================================================TRANSFERENCIA DE MEDICAMENTO A OTRA UNIDAD MEDICA =================
@bp.route('/transferencias')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def listar_transferencias():
    # 1. Mantenemos las tablas viejas por si tienes historial que no quieres perder
    salientes_viejas = TransferenciaSaliente.query.order_by(TransferenciaSaliente.fecha_transferencia.desc()).all()
    entrantes_viejas = TransferenciaEntrante.query.order_by(TransferenciaEntrante.fecha_transferencia.desc()).all()

    # 2. Obtenemos las "Nuevas Transferencias" y "Bajas" desde la tabla de movimientos
    # Filtramos por las palabras clave que pusimos en el proceso masivo
    movimientos_especiales = MovimientoAlmacenFarmacia.query.filter(
        (MovimientoAlmacenFarmacia.observaciones.like('%TRANSFERENCIA_EXTERNA%')) |
        (MovimientoAlmacenFarmacia.observaciones.like('%BAJA_CADUCIDAD%')) |
        (MovimientoAlmacenFarmacia.observaciones.like('%BAJA_MERMA%'))
    ).order_by(MovimientoAlmacenFarmacia.fecha_movimiento.desc()).all()

    return render_template(
        'farmacia/transferencias.html', 
        salientes=salientes_viejas, 
        entrantes=entrantes_viejas,
        especiales=movimientos_especiales # Estos son los nuevos
    )


#============================================================================================================Reporte de inventario================================



@bp.route('/inventario/reporte')
@roles_required(['UsuarioAdministrativo', 'Administrador'])
def reporte_inventario():

    # Cargar observaciones una sola vez
    entradas = EntradaAlmacen.query.with_entities(
        EntradaAlmacen.id_medicamento,
        EntradaAlmacen.lote,
        EntradaAlmacen.observaciones
    ).all()

    obs_map = {
        (e.id_medicamento, e.lote): e.observaciones or ""
        for e in entradas
    }

    # Cargar medicamentos e inventarios
    medicamentos = Medicamento.query.options(
        selectinload(Medicamento.inventario_almacen),
        selectinload(Medicamento.inventario_farmacia)
    ).all()

    reporte = []

    hoy = datetime.utcnow().date()
    limite_vencimiento = hoy + timedelta(days=90)

    def calcular_color(cantidad, min_s, max_s):
        if cantidad <= min_s:
            return "danger"
        if cantidad <= (max_s * 0.5):
            return "warning"
        return "success"

    for med in medicamentos:

        # ==========================
        # ALMACÉN
        # ==========================
        lotes_alm = []
        cant_alm = 0

        for i in med.inventario_almacen:
            if i.cantidad > 0:

                cant_alm += i.cantidad

                obs_texto = obs_map.get(
                    (med.id_medicamento, i.lote),
                    ""
                )

                lotes_alm.append({
                    "lote": i.lote,
                    "cant": i.cantidad,
                    "vence": (
                        i.fecha_vencimiento.strftime('%d/%m/%Y')
                        if i.fecha_vencimiento
                        else 'N/A'
                    ),
                    "observaciones": obs_texto
                })

        # ==========================
        # FARMACIA
        # ==========================
        lotes_far = []
        cant_far = 0

        for i in med.inventario_farmacia:
            if i.cantidad > 0:

                cant_far += i.cantidad

                obs_f = obs_map.get(
                    (med.id_medicamento, i.lote),
                    ""
                )

                lotes_far.append({
                    "lote": i.lote,
                    "cant": i.cantidad,
                    "vence": (
                        i.fecha_vencimiento.strftime('%d/%m/%Y')
                        if i.fecha_vencimiento
                        else 'N/A'
                    ),
                    "observaciones": obs_f
                })

        # ==========================
        # TOTALES
        # ==========================
        total = cant_alm + cant_far

        texto_busqueda_lotes = " ".join(
            [l['lote'] for l in lotes_alm] +
            [l['lote'] for l in lotes_far]
        )

        # ==========================
        # CADUCIDAD
        # ==========================
        tiene_caducados = False
        tiene_proximos = False

        for inv in (
            list(med.inventario_almacen) +
            list(med.inventario_farmacia)
        ):

            if inv.cantidad > 0 and inv.fecha_vencimiento:

                if inv.fecha_vencimiento <= hoy:
                    tiene_caducados = True

                elif inv.fecha_vencimiento <= limite_vencimiento:
                    tiene_proximos = True

        if tiene_caducados:
            estado_caducidad = "ESTADO_CADUCADO"
        elif tiene_proximos:
            estado_caducidad = "ESTADO_PROXIMO"
        else:
            estado_caducidad = "ESTADO_VIGENTE"

        # ==========================
        # REPORTE
        # ==========================
        reporte.append({
            "clave": med.clave,
            "nombre": med.principio_activo,
            "presentacion": med.presentacion,
            "concentracion": med.concentracion,
            "lotes_busqueda": texto_busqueda_lotes,
            "estado_caducidad": estado_caducidad,
            "lotes_almacen": lotes_alm,
            "lotes_farmacia": lotes_far,
            "almacen": {
                "cant": cant_alm,
                "color": calcular_color(
                    cant_alm,
                    med.stock_minimo,
                    med.stock_maximo
                )
            },
            "farmacia": {
                "cant": cant_far,
                "color": calcular_color(
                    cant_far,
                    med.stock_minimo,
                    med.stock_maximo
                )
            },
            "total": {
                "cant": total,
                "color": calcular_color(
                    total,
                    med.stock_minimo,
                    med.stock_maximo
                )
            }
        })

    return render_template(
        'farmacia/reporte_inventario.html',
        reporte=reporte
    )
